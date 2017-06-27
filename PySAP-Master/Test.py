
#!/usr/bin/env Python
# coding=utf-8
"""
Created on Mon Feb 27 12:21:33 2017

@author: rf_hs
"""

import os
import re

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule

from MySql.MySqlDB import MySQL

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

def IsSubString(SubStrList,Str):
    '''
    #判断字符串Str是否包含序列SubStrList中的每一个子字符串
    #>>>SubStrList=['F','EMS','txt']
    #>>>Str='F06925EMS91.txt'
    #>>>IsSubString(SubStrList,Str)#return True (or False)
    '''
    flag=True
    for substr in SubStrList:
        if not(substr in Str):
            flag=False

    return flag
    
def GetFileList(FindPath,FlagStr=[]):
    '''
    #获取目录中指定的文件名
    #>>>FlagStr=['F','EMS','txt'] #要求文件名称中包含这些字符
    #>>>FileList=GetFileList(FindPath,FlagStr) #
    '''
    FileList=[]
    
    FileNames=os.listdir(FindPath)
    
    
    if(len(FileNames)>0):
       
       for fn in FileNames:
                            
           if (len(FlagStr)>0):
               #返回指定类型的文件名
               if (IsSubString(FlagStr,fn)):
                   fullfilename=os.path.join(FindPath,fn)
                   FileList.append(fullfilename)
           else:
               #默认直接返回所有文件名
               fullfilename=os.path.join(FindPath,fn)
               FileList.append(fullfilename)
           
    #对文件名排序
    if (len(FileList)>0):
        FileList.sort()

    return FileList



querysql  = "select count(*) from BzCategory"

getAllLayerSql = """select (count(parent.bz_name)-1) As depth,node.bz_id,node.bz_pid,node.bz_name
                    from bzcategory As node,bzcategory As parent where 
                    node.bz_lft between parent.bz_lft And parent.bz_rgt  And node.bz_isindex=0
                    GROUP BY node.bz_name ORDER BY node.bz_lft"""

truncsql  = "truncate table BzCategory"

deletesql = "delete from BzCategory"

# 批量sql语句
patchsql = """ insert into BzCategory
                      (Bz_Name, Bz_Code, Bz_IndexCode,Bz_Lft,Bz_Rgt,
                       Bz_pid,Bz_isindex) 
                       values(%s,%s,%s,%s,%s,%s,%s)
               """  

#初始化板块指数数据

path = os.path.abspath(os.path.dirname(__file__))  

BInit_FILENAME = 'BoardIndexInit.xlsx'

BInit_FILENAME = os.path.join(path, BInit_FILENAME) 

#获取初始化文件
wb = load_workbook(BInit_FILENAME)

sheet_names = wb.get_sheet_names()

ws = wb.get_sheet_by_name(sheet_names[0])# index为0为第一张表 

ws_row = ws.max_row

ws_col = ws.max_column

collist =[]

sqldatalist =[]

alldata  = tuple(ws.rows)

rowcount = 0

colcount = 0 

BzNum    = 0


#遍历sheet中所有数据
for row in ws.iter_rows():
    
    rowcount = rowcount + 1
    
    if rowcount>1:
        #遍历列中所有元素
        for cell in row:
            colcount = colcount +1 
            
            if colcount>1:
               collist.append(cell.value)
            
        #list转换成为元组
        coltuple = tuple(collist)
        
        #清空list
        collist  = []
        
        sqldatalist.append(coltuple)
    
    colcount = 0


#数据库连接参数  
dbconfig = {'host':'localhost', 
        'port': 3306, 
        'user':'root', 
        'passwd':'1234', 
        'db':'stocksystem', 
        'charset':'utf8'}   

db = MySQL(dbconfig)
  
check_tbsql = "show tables"  
  
db.query(check_tbsql)
    #获取结果列表
results = db.fetchAllRows()

db.query(querysql)

BzNum = db.fetchAllRows()

if BzNum>0:
    db.query(truncsql)
    db.update(deletesql)

# 插入初始化板块指数数据

db.insertmany(patchsql,sqldatalist)

'''
    整个体系分为多层，板块指数->（行业，概念，风格，地域）->各细分类。
    
    1. 获取所有分类的层次，取出节点深度为1的分类；
    2. 查询所有分类为1的板块指数文件夹目录，并读取其中所有.txt文档，根据文档名称找到相应节点；
    3. 将文档中指数插入到分类表中。
    
'''
db.query(getAllLayerSql)

BzLayerResults =db.fetchAllRows() 

qrowcount = 0

qcolcount = 0 

LayerDict = {}

NameDict  = {}

# 板块指数1级分类 数据库中暂时有4大类：行业，概念，风格，地域
BzlayerList = []

# 板块指数2级分类，1级分类中的子类
BzlayerList_2 = []

NameList= []

allBzstr =""

tmpstr  =""

lentmpstr = 0

bz_layerid =0 

bz_id = 0 

bz_pid =0 

bz_xpid = 0

bz_name=""

bz_path=''
# 合并所有节点

for bzrow in BzLayerResults:
    
    #tuple 转str
    
    bzstr  = bzrow.__str__()
    
    #第一个设置为层次,求出板块指数层次
    
    #第二个层次生成，以名称为索引的dict
    
    tmpstr   = bzstr.replace("(","")
    
    tmpstr   = tmpstr.replace(")","")
    
    NameList = tmpstr.split(",")
    
    if len(NameList)==4: 
        
       bz_layerid    = NameList[0]
       
       bz_id         = NameList[1]
       
       bz_pid        = NameList[2]
       
       bz_name       = NameList[3]
       
       bz_name       = bz_name.strip()
       
       #按照父节点，子节点，层次节点区分
            
       tmpstr  =  bz_id +"," + bz_pid +","+bz_layerid
       
       NameDict[bz_name] = tmpstr
    
    
    bzpos  = bzstr.find(",")
    
    if bzpos!=-1:
        
        bzlayer = bzstr[1:bzpos]
         
        if allBzstr=="":
            
            LayerDict[bzlayer] =   bzstr          
            
            allBzstr = "|"+bzlayer+"|"
        else:
            if allBzstr.find(bzlayer)==-1:
                allBzstr = allBzstr + bzlayer +"|"
                LayerDict[bzlayer] =   bzstr  
            else:
                tmpstr = LayerDict[bzlayer]
                LayerDict[bzlayer] =   bzstr +tmpstr

#遍历dict，处理dict中的字符数据用|进行隔断
for layerkey in LayerDict:
    tmpstr = LayerDict[layerkey]
    tmpstr = tmpstr.replace("(","")
    tmpstr = tmpstr.replace(")","|")
    tmpstr = tmpstr[:len(tmpstr)-1]
    
    LayerDict[layerkey] = tmpstr


FlagStr=['txt']

fileList=[]

tmpPath = "BoardIndex\\config\\"

boardPath = os.path.join(path, tmpPath)


bz_lists =[]

layerlists =[]

layercount = 0


#读取节点为1 层次的所有分类

layerstr = LayerDict["1L"]

layerstr_xf = LayerDict["3L"]

#layerstr   = layerlists[1].__str__()

bz_lists   = layerstr.split("|")

flie_list  = []

flag       = 0 

# 读出所有文件夹下面的行业配置文件

for bzlist in bz_lists:
    
    BzlayerList = bzlist.split(",")
            
    layercount  = 0
            
    #初始化板块指数节点
            
    for bzlayer in BzlayerList:
                
        if layercount ==0:  # 层次节点
            bz_layerid =  bzlayer
                    
        if layercount ==1:  # 板块节点
            bz_id  = bzlayer
                
        if layercount ==2:  # 板块父节点
            bz_pid = bzlayer
                
        if layercount ==3:  # 节点名称
            bz_name = bzlayer
                
        layercount = layercount +1
            
    bz_name = bz_name.decode("unicode_escape")    
            
    bz_name = bz_name.encode("UTF-8")
            
    bz_name = bz_name.replace("u'","")
            
    bz_name = bz_name.replace("'","").strip()
            
    bz_path = os.path.join(boardPath, bz_name)
            
    fileList = GetFileList(bz_path,FlagStr)
            
    keys      =  NameDict.keys()
            
    #插入文件内的数据
             
    for flist in fileList:
                 
        flie_list =  flist.split("\\")
                 
        bz_name   =  flie_list[len(flie_list)-1]
                 
        bz_name   =  bz_name.replace(".txt","").strip()
                 
        bz_name   =  repr(bz_name) 
        
        splines   = ""
                              
        try:
                     
            rfile = open(flist,'r')
                     
            #读取第一行内容
            Fline = rfile.readline() 
                     
            Fstrlist = Fline.split()
                 
            Flen = len(Fstrlist)
                     
            # 数量不多一次性全部取出
            lines = rfile.readlines()
                     
            splines = "|".join(lines)
                         
            splines = "|"+splines+"|"
                         
            #print splines
                          
            #re.sub(r"\s{2,}", " ", splines)
                          
            splines = re.sub(r"\s+", ",", splines)
                     
            print splines
                                         
            #如果长度大于２，该分支有多级分类
                        
            if Flen<=2:
                         
            # 转换成特定字符串，存入存储过程，建立分类表格   
                   
                mm = 1
                         
                flag = 0 
                   
            #单分支板块指数，在存储过程中直接插入
                        
            #splines =','.join(filter(lambda x: x, splines.split(' ')))
                       
            else:
                # 此处设计一级与二级两大块的入库
                # 暂时只分到二级行业，这里可以留出做扩展用
                         
                flag  = 1
                                                 
                xfstr = Fstrlist[Flen-1]
                         
                xfstr = xfstr.decode()
                         
                xfstr = repr(xfstr)
                         
                #xfstr = xfstr.decode("unicode_escape")    
            
                # xfstr = xfstr.encode("UTF-8")
                         
                # 需求出细分行业的父ID
                         
                tmpstr  = NameDict[xfstr] 
                         
                tmplsit = tmpstr.split(",")
                         
                bz_xpid = long(tmplsit[0])
                         
                # 传入bz_id,bz_xpid,标识三个参数
                   
                kk = 1
                   #统一存储过程参数配置，1.行业类别字符串；2.单级多级标识；3. 单级父节点；4.多级父节点；
                     
                     
            if NameDict.has_key(bz_name):
                     
                tmpstr    =  NameDict[bz_name] 
                         
                tmplsit   =  tmpstr.split(",")
                                 
                bz_pid    =  long(tmplsit[0])
                         
                params    =(int(bz_pid),int(bz_xpid),flag,splines)
                         
                db.call_procedure_sp('sp_boardIndex_bulid',params)
                       
                     # 统一调用存储过程，参数（flag，bz_pid,bz_xpid,tmp）
                     
                     
        finally:
            if rfile:
               rfile.close() 
            
            

m =1

'''
for layerkey in LayerDict:
    
    #从第2层节点开始处理
    
    
    if layerkey>=2:
        
        tmpstr   = LayerDict[layerkey]
        
        bz_lists = tmpstr.split("|")
        
        for bzlist in bz_lists:
            
            BzlayerList = bzlist.split(",")
            
            layercount  = 0
            
            #初始化板块指数节点
            
            for bzlayer in BzlayerList:
                
                layercount = layercount +1
                
                if layercount ==1:  # 层次节点
                   bz_layerid =  bzlayer
                    
                if layercount ==2:  # 板块节点
                   bz_id  = bzlayer
                
                if layercount ==3:  # 板块父节点
                   bz_pid = bzlayer
                
                if layercount ==4:  # 节点名称
                   bz_name = bzlayer
            
                        
            
            bz_path = os.path.join(boardPath, bz_name)
            
            fileList = GetFileList(bz_path,)
                
                
            
            
     
        
    
    
     
     for bztuple in bzlist:
         
         qcolcount = qcolcount +1
     
         if qcolcount==1:
             bz_id = bztuple
         
         if qcolcount==2:
             
             bz_name = bztuple
             
             bz_path = os.path.join(boardPath, bz_name)
 
             fileList = GetFileList(bz_path,FlagStr)
             
             #插入文件内的数据
             
             for flist in fileList:
                 
                 try:
                     
                     rfile = open(flist,'r')
                     
                     #读取第一行内容
                     Fline = rfile.readline() 
                     
                     
                     Fstrlist = Fline.split()
                 
                     Flen = len(Fstrlist)
                     
                     # 数量不多一次性全部取出
                     lines = rfile.readlines()
                                             
                     #如果长度大于２，该分支有多级分类
                     
                     if Flen<=2:
                         
                         # 转换成特定字符串，存入存储过程，建立分类表格                       
                         splines = "|".join(lines)
                         
                         splines = "|"+splines+"|"
                         
                         
                         
                         print splines
                          
                         #re.sub(r"\s{2,}", " ", splines)
                          
                         tmp = re.sub(r"\s+", ",", splines)
                         
                         print tmp
                         
                         #splines =','.join(filter(lambda x: x, splines.split(' ')))
                        
                         
                         
                     else:
                         kk = 1
                         
                     
                     k=1
                 finally:
                     if rfile:
                         rfile.close() 
  
             
             
     
     qcolcount = 0
     
     





 
sqldatalist=[]


db.close() 
'''
        


