# -*- coding: utf-8 -*-
"""
Created on Tue Mar 14 11:46:55 2017

@author: rf_hs
"""


#!/usr/bin/env Python
# coding=utf-8
"""
Created on Mon Feb 27 12:21:33 2017

@author: rf_hs
"""

import os
import re

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule

from FileStringHandle import FileStringHandle

from FileStringHandle import loop_file

import sys

class Board:
      
      def __init__(self):
          
          # 查询板块指数内容个数
                            
          self.querysql  = "select count(*) from BzCategory"
          
          
          self.truncsql  = "truncate table BzCategory"
          
          self.deletesql = "delete from BzCategory"
          
          
          # 查询板块指数层次个数  
          self.getLayerSql = """select (count(parent.bz_name)-1) As depth,node.bz_id,node.bz_pid,node.bz_name
                                    from bzcategory As node,bzcategory As parent where 
                                    node.bz_lft between parent.bz_lft And parent.bz_rgt  And node.bz_isindex=0
                                    GROUP BY node.bz_name ORDER BY node.bz_lft"""
          
          # 查询板块信息
          
          self.queryboardinfosql  = "select count(*) from boardindexbaseinfo"
          
          self.truncboardinfosql  = "truncate table boardindexbaseinfo"
          
          self.deleteboardinfosql = "delete from boardindexbaseinfo"
          
          
          self.boardRelatequerysql  = "select count(*) from boardstock_related"
          
          self.truncboardRelatesql = "truncate table boardstock_related"
          
          self.deleteboardRelatesql = "delete from boardstock_related"
          
          
          # 板块节点批量插入初始化
          self.patchbkstocksql = """ insert into boardstock_related
                              (board_id, board_name, stock_id,stock_name) 
                               values(%s,%s,%s,%s)
                          """
                          
         # 板块-个股关联插入
          self.patchBZCatesql = """ insert into BzCategory
                              (Bz_Name, Bz_Code, Bz_IndexCode,Bz_Lft,Bz_Rgt,
                               Bz_pid,Bz_isindex) 
                               values(%s,%s,%s,%s,%s,%s,%s)
                          """ 
          
          # 个股信息批量插入初始化
          self.patchstockinfosql = """ insert into stockbaseinfo
                                  (stock_code, stock_name, stock_InMarketDate,stock_BelongMarket,
                                   stock_MarketSymbol,stock_MarketBlocks) 
                                   values(%s,%s,%s,%s,%s,%s)
                                   """
          # 查询节点对应  
          self.allbkcode         =""" select bkindexnode.bz_name,bkindexnode.bz_indexcode from bzcategory As bkindexnode where
                                      bkindexnode.bz_isindex=1 
                            
                                  """      
         
          # 查询个股权重
          
          self.queryweightsql  = "select count(*) from stock_weight"
          
          self.truncweightsql  = "truncate table stock_weight"
          
          self.deleteweightsql = "delete from stock_weight"
          
          
          self.patchweightsql = """ insert into stock_weight
                                (stock_id, stock_name, stock_weight) 
                                 values(%s,%s,%s)
                                """ 
                    
          
          # 板块节点路径
          self.BInit_FILENAME = 'BoardIndexInit.xlsx' 
          
          #配置路径
          self.configpath     = "BoardIndex\\config\\"
          
          # A股所有股票路径
          
          self.allstockpath   = "Data\\History\\BoardIndex\\沪深A股数据\\"
          
          self.bkstockpath    = "\\Data\\History\\BoardIndex\\板块个股组成"
         
          
          self.path = os.path.abspath(os.path.dirname(__file__)) 
           
          reload(sys)
          
          sys.setdefaultencoding('utf-8')
      
      #初始化板块指数数据
      def boardLayerInit(self,mysqldb): 
          
          
          
            filename = os.path.join(self.path, self.BInit_FILENAME) 

            #获取初始化文件

            wb = load_workbook(filename)

            sheet_names = wb.get_sheet_names()

            ws = wb.get_sheet_by_name(sheet_names[0])# index为0为第一张表 

            collist =[]

            sqldatalist =[]
            
            rowcount = 0

            colcount = 0 

            BzNum    = 0
            
            #遍历sheet中所有数据
            for row in ws.iter_rows():
                
                rowcount = rowcount + 1
                
                if rowcount>1:
                    #遍历列中所有元素
                    for cell in row:
                        colcount = colcount +1 
                        
                        if colcount>1:
                           collist.append(cell.value)
                        
                    #list转换成为元组
                    coltuple = tuple(collist)
                    
                    #清空list
                    collist  = []
                    
                    sqldatalist.append(coltuple)
                
                colcount = 0
            
            #获取结果列表
            mysqldb.query(self.querysql)

            BzNum = mysqldb.fetchAllRows()
            
            # 初始化板块节点，同时更新
            if BzNum>0:
                mysqldb.query(self.truncsql)
                mysqldb.update(self.deletesql)
                
                mysqldb.query(self.truncboardinfosql)
                mysqldb.update(self.deleteboardinfosql)

            # 插入初始化板块指数数据
            
            mysqldb.insertmany(self.patchBZCatesql,sqldatalist)
            
            sqldatalist = []


      def boardLayerInsert(self,mysqldb):
          
          '''
              整个体系分为多层，板块指数->（行业，概念，风格，地域）->各细分类。
    
                1. 获取所有分类的层次，取出节点深度为1的分类；
                2. 查询所有分类为1的板块指数文件夹目录，并读取其中所有.txt文档，根据文档名称找到相应节点；
                3. 将文档中指数插入到分类表中。
    
          '''
          
          fsh = FileStringHandle()          
          
          mysqldb.query(self.getLayerSql)

          BzLayerResults =mysqldb.fetchAllRows() 
          
          LayerDict = {}

          NameDict  = {}

          # 板块指数1级分类 数据库中暂时有4大类：行业，概念，风格，地域
          BzlayerList = []

          NameList= []

          allBzstr =""

          tmpstr  =""
        
          bz_layerid =0 
        
          bz_id = 0 
        
          bz_pid =0 
        
          bz_xpid = 0
        
          bz_name=""
        
          bz_path=''
          
          # 合并所有节点
        
          for bzrow in BzLayerResults:
            
              #tuple 转str
            
              bzstr  = bzrow.__str__()
            
              #第一个设置为层次,求出板块指数层次
            
              #第二个层次生成，以名称为索引的dict
            
              tmpstr   = bzstr.replace("(","")
            
              tmpstr   = tmpstr.replace(")","")
            
              NameList = tmpstr.split(",")
            
              if len(NameList)==4: 
                
                   bz_layerid    = NameList[0]
                   
                   bz_id         = NameList[1]
                   
                   bz_pid        = NameList[2]
                   
                   bz_name       = NameList[3]
                   
                   bz_name       = bz_name.strip()
                   
                   #按照父节点，子节点，层次节点区分
                        
                   tmpstr  =  bz_id +"," + bz_pid +","+bz_layerid
                   
                   NameDict[bz_name] = tmpstr
            
            
              bzpos  = bzstr.find(",")
            
              if bzpos!=-1:
                
                  bzlayer = bzstr[1:bzpos]
                 
                  if allBzstr=="":
                    
                      LayerDict[bzlayer] =   bzstr          
                    
                      allBzstr = "|"+bzlayer+"|"
                  else:
                      if allBzstr.find(bzlayer)==-1:
                          allBzstr = allBzstr + bzlayer +"|"
                          LayerDict[bzlayer] =   bzstr  
                      else:
                          tmpstr = LayerDict[bzlayer]
                          LayerDict[bzlayer] =   bzstr +tmpstr
        
          #遍历dict，处理dict中的字符数据用|进行隔断
          for layerkey in LayerDict:
                tmpstr = LayerDict[layerkey]
                tmpstr = tmpstr.replace("(","")
                tmpstr = tmpstr.replace(")","|")
                tmpstr = tmpstr[:len(tmpstr)-1]
                
                LayerDict[layerkey] = tmpstr
        
          FlagStr=['txt']
        
          fileList=[]
          
          #testpath = os.getcwd()
        
          boardPath = os.path.join(os.getcwd(), self.configpath)
         
          bz_lists =[]
        
          layercount = 0
                
          #读取节点为1 层次的所有分类
        
          layerstr = LayerDict["1L"]
        
          #layerstr   = layerlists[1].__str__()
        
          bz_lists   = layerstr.split("|")
        
          flie_list  = []
        
          flag       = 0 
        
          # 读出所有文件夹下面的行业配置文件
        
          for bzlist in bz_lists:
            
                BzlayerList = bzlist.split(",")
                        
                layercount  = 0
                        
                #初始化板块指数节点
                        
                for bzlayer in BzlayerList:
                            
                    if layercount ==0:  # 层次节点
                        bz_layerid =  bzlayer
                                
                    if layercount ==1:  # 板块节点
                        bz_id  = bzlayer
                            
                    if layercount ==2:  # 板块父节点
                        bz_pid = bzlayer
                            
                    if layercount ==3:  # 节点名称
                        bz_name = bzlayer
                            
                    layercount = layercount +1
                        
                bz_name = bz_name.decode("unicode_escape")    
                        
                bz_name = bz_name.encode("UTF-8")
                        
                bz_name = bz_name.replace("u'","")
                        
                bz_name = bz_name.replace("'","").strip()
                        
                bz_path = os.path.join(boardPath, bz_name)
                
                        
                fileList = fsh.GetFileList(bz_path.decode('utf-8'),FlagStr)
                        
                #插入文件内的数据
                         
                for flist in fileList:
                    
                    bz_xpid   =  0
                             
                    flie_list =  flist.split("\\")
                             
                    bz_name   =  flie_list[len(flie_list)-1]
                             
                    bz_name   =  bz_name.replace(".txt","").strip()
                             
                    bz_name   =  repr(bz_name) 
                    
                    splines   =  ""
                                          
                    try:
                                 
                        rfile = open(flist,'r')
                                 
                        #读取第一行内容
                        Fline = rfile.readline() 
                                 
                        Fstrlist = Fline.split()
                             
                        Flen = len(Fstrlist)
                                 
                        # 数量不多一次性全部取出
                        lines = rfile.readlines()
                                                                                  
                        #如果长度大于２，该分支有多级分类
                                    
                        if Flen<=2:
                                     
                        # 转换成特定字符串，存入存储过程，建立分类表格   
                               
                            flag = 0 
                            
                                      
                            splines = "|".join(lines)
                                         
                            splines = "|"+splines+"|"
                                         
                            #print splines
                                          
                            #re.sub(r"\s{2,}", " ", splines)
                                          
                            splines = re.sub(r"\s+", ",", splines)
                            
                            splines = splines.replace(",|","@")
                            
                               
                        #单分支板块指数，在存储过程中直接插入
                                    
                        #splines =','.join(filter(lambda x: x, splines.split(' ')))
                                   
                        else:
                            # 此处设计一级与二级两大块的入库
                            # 暂时只分到二级行业，这里可以留出做扩展用
                                     
                            flag  = 1
                            
                            xfstr = Fstrlist[Flen-1]
                                     
                            xfstr = xfstr.decode()
                                     
                            xfstr = repr(xfstr)
                                     
                            #xfstr = xfstr.decode("unicode_escape")    
                        
                            # xfstr = xfstr.encode("UTF-8")
                                     
                            # 需求出细分行业的父ID
                                     
                            tmpstr  = NameDict[xfstr] 
                                     
                            tmplsit = tmpstr.split(",")
                                     
                            bz_xpid = long(tmplsit[0])
                                     
                            # 传入bz_id,bz_xpid,标识三个参数
                               
                            #统一存储过程参数配置，1.行业类别字符串；2.单级多级标识；3. 单级父节点；4.多级父节点；                                 
                                 
                        if NameDict.has_key(bz_name):
                                 
                            tmpstr    =  NameDict[bz_name] 
                                     
                            tmplsit   =  tmpstr.split(",")
                                             
                            bz_pid    =  long(tmplsit[0])
                            
                            if flag:
                                
                                linecount = 0
                            
                                for line in lines :
                                    
                                    linelist = line.split()
                                     
                                    #strpos = line.find("1")
                                    
                                    linelen = len(linelist)
                                    
                                    if linelen==4:
                                        
                                        if linelist[2]=="1" and linelist[3]=="0":
                                           tmpstr = linelist[1]
                                           bkname  = BzLayerResults[int(bz_pid)-1]
                                           
                                           bklen   =  len(bkname)
                                
                                           bkstr   =  bkname[bklen-1]
                                           
                                           line    = line.replace(tmpstr,bkstr+"-"+tmpstr)                                       
                                           
                                        
                                        if linelist[2]=="0" and linelist[3]=="2":
                                           tmpstr = linelist[1]
                                           
                                           bkname  = BzLayerResults[int(bz_xpid)-1]
                                           
                                           bklen   =  len(bkname)
                                            
                                           bkstr   =  bkname[bklen-1]
                                           
                                           line    = line.replace(tmpstr,bkstr+"-"+tmpstr)      
                                        
                                        lines[linecount] = line
                                        
                                    linecount = linecount +1
                                
                                
                                splines = "|".join(lines)
                                             
                                splines = "|"+splines+"|"
                                             
                                #print splines
                                              
                                #re.sub(r"\s{2,}", " ", splines)
                                              
                                splines = re.sub(r"\s+", ",", splines)
                                
                            else:
                               
                                bkname  = BzLayerResults[int(bz_pid)-1]
                                
                                bklen   =  len(bkname)
                                
                                bkstr   =  bkname[bklen-1]
                                
                                splines = splines.replace(",",","+bkstr+"-") 
                                
                                splines = splines.replace("@",",|") 
                                
                                
                            print splines
                                     
                            params    =  (int(bz_pid),int(bz_xpid),flag,splines)
                                     
                            mysqldb.call_procedure_sp('sp_boardIndex_bulid',params)
                                   
                            # 统一调用存储过程，参数（flag，bz_pid,bz_xpid,tmp）                                 
                                 
                    finally:
                        if rfile:
                           rfile.close()
                           
      def dealstockfile(self,dealtype,flist,mysqldb):
          
          # dealtype为0  代码 名称 细分行业 地区 上市日期
          # 注意文本编码问题，不然读取数据会很头疼
          
          stockcode     =''
          
          stockname     =''
          
          stockline     =''
          
          stockarea     =''
          
          stockinmarket =''
          
          #附加插入字段
          stockbelongmarket =''
          
          stockmarketsymbol =''
          
          stockmarketblocks  =''
          
          stocklist         =[]
          
          tmpstr            =''
          
          stocksqldata      =()
          
          linedict          ={}
          
          linekey           =''
          
          areadict          ={}
          
          areakey           =''
          
          startpos          = 0
          
          endpos            = 0
          
          prestartpos       = 0
                
          commstr           =''
          
          headstr           ='代码    名称|'
          
          try:
            
            filepath = flist.__str__()
            
            filepath  = unicode(filepath)
           
            rfile = open(filepath,'r')
                                 
            #读取第一行内容
            Fline = rfile.readline() 
                                 
            Fstrlist = Fline.split()
                             
            Flen = len(Fstrlist)
            
             # 数量不多一次性全部取出
            lines = rfile.readlines()
            
            if int(dealtype)==0:
                    
                splines = "|".join(lines)
                                     
                splines = "|"+splines+"|"
                                      
                splines = re.sub(r"\s+", ",", splines)
                                 
                print splines
                
                if Flen==5:
                    
                    endpos          = len(splines)
                                           
                    while startpos!=-1:
                        
                        prestartpos     = startpos
                        
                        startpos        = splines.find("|",startpos+1,endpos)
                        
                        tmpstr          = splines[prestartpos+1:startpos-1]
                        
                        strlist         = tmpstr.split(",")
                        
                        if len(strlist)==5:
                                                
                            stockcode       = strlist[0]
                            
                            stockname       = strlist[1]
                            
                            stockline       = strlist[2]
                            
                            stockarea       = strlist[3]
                            
                            stockinmarket   = strlist[4]                       
                        
                            # 600开头为上海主板，000001到002000之间深圳主板，002001到300000中小板，300开头创业板
                            
                            # 对name进行处理
                            
                            if int(stockcode)>600000:
                                stockbelongmarket ='01'
                                stockmarketsymbol = 'SH'
                                stockmarketblocks ='011'
                            else:
                                stockbelongmarket ='02'
                                stockmarketsymbol = 'SZ'
                                
                                if int(stockcode)>300000:
                                    stockmarketblocks ='023'
                                elif int(stockcode)>0 and int(stockcode)<=002000:
                                    stockmarketblocks ='022'                               
                                else:
                                    stockmarketblocks ='021'                                     
                        
                            # 插入股票信息
                            stocksqldata =(stockcode,stockname,stockinmarket,stockbelongmarket,stockmarketsymbol,stockmarketblocks)
                            
                            if stockinmarket!="--":
                                stocklist.append(stocksqldata)
                            
                            # 先取地域类股票 
                                                    
                            commstr    = stockcode +","+stockname+"|"
                            
                            areakey    = stockarea.decode("UTF-8")
                            
                            linekey    = stockline.decode("UTF-8")
                                                
                            # 生成地域行业文件                                               
                            
                            if areadict.has_key(areakey):
                                
                                tmpstr    =  areadict[areakey]   
                               
                                tmpstr    =  tmpstr + commstr
                               
                                areadict[areakey] = tmpstr
                               
                            else:
                                
                                areadict[areakey] =   commstr 
                             
                            # 生成细分行业文件
                            
                            if linedict.has_key(linekey):
                                
                                tmpstr    =  linedict[linekey]   
                               
                                tmpstr    =  tmpstr + commstr
                               
                                linedict[linekey] = tmpstr
                               
                            else:
                                
                                linedict[linekey] =   commstr
                    
                    #  股票信息数据入库                  
                    
                    #sqlstr = self.patchstockinfosql                     
                    
                    mysqldb.insertmany(self.patchstockinfosql,stocklist)
                    
                    stocklist = []
                                                          
                    # 所有数据读取完毕，准备写文件
                    # 行业目录 PySAP-Master\Pycode\SAP\Data\History\BoardIndex\板块个股组成\行业\通达信\二级
                    # 地域目录 PySAP-Master\Pycode\SAP\Data\History\BoardIndex\板块个股组成\地域
                    
                    linedir  = "\\Data\\History\\BoardIndex\\板块个股组成\\行业\\通达信\\细分行业\\"

                    areadir  = "\\Data\\History\\BoardIndex\\板块个股组成\\地域\\通达信地域\\"                    
                    
                                                       
                    #处理地域
                    for edict in areadict:
                        
                        if edict.__str__()!="黑龙江":
                            tmpstr = unicode(areadir)+unicode(edict.__str__()+"板块")+".txt"
                        else:
                            tmpstr = unicode(areadir)+unicode(edict.__str__())+".txt"
                            
                            
                        fpath  = os.getcwd()+tmpstr
                                           
                        wfile  = open(fpath, 'w')
                        
                        # 还需处理字典中的字符串 

                        contentstr =unicode(headstr)+areadict[edict].__str__()
                        
                        contentstr = contentstr.replace('|','\r\n')
                        
                        wfile.writelines(contentstr)
                        
                        wfile.close()
                        
                    #处理行业
                    for edict in linedict:
                        
                        tmpstr = unicode(linedir)+unicode(edict.__str__())+".txt"
                        
                        fpath = os.getcwd()+ tmpstr
                
                        
                        wfile = open(fpath, 'w')
                        
                        # 还需处理字典中的字符串 
                        
                        contentstr =unicode(headstr)+ linedict[edict].__str__()
                        
                        contentstr = contentstr.replace('|','\r\n')
                        
                        wfile.writelines(contentstr)
                        
                        wfile.close()
            # 处理股本结构
            if int(dealtype)==1:                
                pass;
            
            # 处理股票权重（暂时按中证全指分类）
            if int(dealtype)==2:
                
               weightlist =[]
               
               if Flen==4:
                  
                  for line in lines:
                      
                      linelist  = line.split()
                      
                      if len(linelist)==4:
                          
                          stockid     = linelist[1]
                          
                          stockname   = linelist[2]
                          
                          stockweight = linelist[3]
                
                          weightTuple =(stockid,stockname,stockweight)
                          
                          weightlist.append(weightTuple)
                          
                  
                  #获取结果列表
                  mysqldb.query(self.queryweightsql)
        
                  mysqldb.query(self.truncweightsql)
                        
                  mysqldb.update(self.deleteweightsql)
                  
                  mysqldb.insertmany(self.patchweightsql,weightlist)         
              
          finally:
              if rfile:
                 rfile.close()
                 

      def boardStockDeal(self,mysqldb):
          
          ''' 读取沪深A股数据，日后可扩展的处理
              0-沪深Ａ股
              1-股本结构
              2-股东结构
              3-板块信息
          '''
          
          # 股票板块路径
          
          bsp = FileStringHandle()      
          
          # 读取A股所有股票数据并对信息入库，
          aspath  = os.path.join(os.getcwd(), self.allstockpath)
          
          FlagStr=['.txt']      
          
          dealtype = -1 
          
          fileList = bsp.GetFileList(aspath.decode('utf-8'),FlagStr)   
          
          for flist in fileList:
              
              pos = flist.find("-rfhs-")
              
              if pos!=-1:
                  dealtype = flist[pos-2:pos]
                  
              # 处理股票基本信息，板块基本信息，股票股本信息，股票股东信息等
              if int(dealtype)!=-1:
                  self.dealstockfile(dealtype,flist,mysqldb)
          
      def boardStockLink(self,mysqldb):
         
          # 读取关联板块信息
          # 遍历文件夹中所有文件
          
          bktuple =()
          #获取结果列表
          mysqldb.query(self.allbkcode)
          
          bklist = mysqldb.fetchAllRows()
          
          bkdict ={}
          
          bklinelist =[]
          
          for bkrow in bklist:
              
              bkstr  = bkrow.__str__()
              
              tmplist = bkstr.split(",")
              
              if len(tmplist)==2:
                  bkey    = tmplist[0].replace("(","")
                  bkvalue = tmplist[1].replace("L)","")
              
              if not bkdict.has_key(bkey):
                  bkdict[bkey]  = bkvalue
              
          root_dir  = os.getcwd()+self.bkstockpath
          
          cmplist   = self.bkstockpath.split("\\")
                  
          cmpstr    = cmplist[len(cmplist)-1]
          
          short_exclude = []                            ###不包含检查的短目录、文件  
          
          long_exclude  = []                           ###不包含检查的长目录、文件  
    
          file_extend   = ['.txt']                     ###包含检查的文件类型  
    
          lf = loop_file(root_dir.decode('utf-8'), short_exclude, long_exclude, file_extend)  
          
    
          for f in lf.start(lambda f: f): 
              
              pos    = f.find(cmpstr)

              strlen = len(f)
              
              lfstr  = f[pos:strlen]
              
              lflist = lfstr.split("\\")

              bklen  = len(lflist)
              
              if bklen==4: #一层行业结构，目录之上为总目录              
                 fbkname   = lflist[2]
                 bkname    = lflist[3]
                      
            
              if bklen==5: # 二层行业结构              
                 fbkname = lflist[3]
                 bkname  = lflist[4]
             
              if bklen==6: # 三层行业结构           
                 fbkname = lflist[4]
                 bkname  = lflist[5]
              
              
              bknamepos = bkname.find(".txt")
                 
              bkname    = bkname[0:bknamepos]
              
              timepos   = bkname.find("201")
              
              if timepos==-1:
                 bkkey     = fbkname +"-"+bkname
              else:
                 bkkey     = fbkname +"-"+bkname[0:timepos] 
              
              bkstr        = bkkey
              
              bkkey        = repr(bkkey)
              
              
              if bkdict.has_key(bkkey):
                  
                 bkcode    = bkdict[bkkey]
                 
                 try:
                     rfile = open(f,'r')
                     
                     #读取第一行内容
                     Fline = rfile.readline() 
                     
                     Fstrlist = Fline.split()
                                 
                     Flen = len(Fstrlist)
                                     
                     # 数量不多一次性全部取出
                     lines = rfile.readlines()
                     
                     stockname =""
                     
                     stockcode =""
         
                     
                     #标准结构 直接读取（通达信中地域与细分行业是该结构）                     
                     if Flen>=2:
                         
                         for line in lines:
                             
                             # 处理字符串，分逗号分隔，空格分隔两种，
                             # 空格分隔可能出现 多条数据
                             
                             # 逗号是细分行业处理格式
                             
                             
                             
                             
                             if line.find(",")==-1:
                                
                                # 非逗号处理格式，查找是否为多数据格式（以.号分隔）
                                strpos  = line.find(".")
                                
                                #非逗号分隔 即为空格分隔 申万二级行业
                                
                                if strpos==-1:
                                    linelist = line.split()
                                    
                                    if len(linelist)>=2:
                                       stockname = linelist[1]
                                       
                                       stockcode = linelist[0]
                                       
                                       bktuple = (bkcode,bkstr,stockcode,stockname) 
                                
                                       bklinelist.append(bktuple)
                                else:
                                    #取出前2条数据
                                    if strpos>2:
                                       line     = line[0:strpos-2]
                                     
                                    linelist = line.split()
                                    
                                    if len(linelist)>=2:
                                       stockcode = linelist[0]
                                       stockpos  = len(stockcode)    
                                       stockname = line[stockpos+1:]     
                                       
                                       stockname = ''.join(stockname.split())
                                     
                                                 
                                       bktuple = (bkcode,bkstr,stockcode,stockname) 
                                
                                       bklinelist.append(bktuple)
                             else:
                                linelist = line.split(",") 
                                
                                if len(linelist)>=2:
                                    stockname = linelist[1]
                                    stockcode = linelist[0]
                                    bktuple = (bkcode,bkstr,stockcode,stockname) 
                                    bklinelist.append(bktuple)
                     
                     #bktuple = ()
                     
                 finally:
                    
                    if rfile:
                      rfile.close()
              
                 
          # 录入完毕 准备入库
                 
          #获取结果列表
          mysqldb.query(self.boardRelatequerysql)

          bsrNum = mysqldb.fetchAllRows()

          if bsrNum>0:
             mysqldb.query(self.truncboardRelatesql)
             mysqldb.update(self.deleteboardRelatesql)
             
          mysqldb.insertmany(self.patchbkstocksql,bklinelist)
          
          bklinelist = []
          
    
           
          
          
          
                      
                  
                  
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
          
